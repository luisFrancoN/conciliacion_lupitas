#Programa para conciliacion de depositos realizado por José Luis Franco Navarro 14/14/2021
from openpyxl import *
import sqlite3
from os import system
import time

import os

'''wd = os.getcwd()
print("working directory is ", wd)

filePath = __file__
print("This script file path is ", filePath)

absFilePath = os.path.abspath(__file__)
print("This script absolute path is ", absFilePath)

path, filename = os.path.split(absFilePath)
print("Script file path is {}, filename is {}".format(path, filename))'''

'''def hide_file(nameDB):
    import win32file
    import win32con
    flags = win32file.GetFileAttributesW(nameDB)
    win32file.SetFileAttributes(nameDB, 
        win32con.FILE_ATTRIBUTE_HIDDEN | flags)'''

def borrarDeBD(cur, idArchivo):
    cur.execute("DELETE FROM envios WHERE Id_documento = '"+str(idArchivo[0])+"'")
    cur.execute("DELETE FROM banco WHERE Id_documento = '"+str(idArchivo[0])+"'")
    cur.execute("DELETE FROM acreditaciones WHERE Id_documento = '"+str(idArchivo[0])+"'")
    cur.execute("DELETE FROM netpay WHERE Id_documento = '"+str(idArchivo[0])+"'")
    cur.execute("DELETE FROM archivos WHERE ID = '"+str(idArchivo[0])+"'")

def crearArchivo(cur, nomHojas, nombreArchivo, idArchivo, ingresado):
    if ingresado:
        archivo = Workbook()
        
        hoja = archivo.active
        hoja.title = nomHojas[0]
            
        for i in nomHojas:
            if i != hoja.title:
                hoja = archivo.create_sheet(i)
        
        hoja = archivo[nomHojas[0]]        
        
        cur.execute("SELECT F_trabajo, Fecha_alta, Sucursal, Caja, Turno, ID_envio, Sucursal_ref, Caja_ref, Tipo, Concepto, Detalle_concepto, ID_ref, Estado, Total, Referencia, Mov_banco, Mov_suc, Conciliacion FROM envios WHERE Id_documento = '"+str(idArchivo[0])+"'")#Encontrado = 'No'")
    
        envios = cur.fetchall()
                        
        envios.insert(0, ("F_trabajo", "Fecha_alta", "Sucursal", "Caja", "Turno", "ID_envio", "Sucursal_ref", "Caja_ref", 
                    "Tipo", "Concepto", "Detalle_concepto", "ID_ref", "Estado", "Total", "Referencia", "Mov_banco", "Mov_suc", "Conciliado"))
        for datos in envios:
            hoja.append(datos)
        
        hoja = archivo[nomHojas[1]]
            
        cur.execute("SELECT Cuenta, Fecha, Referencia_num, Importe, Transaccion, Leyenda_1, Leyenda_2, info_adicional, Conciliacion FROM banco WHERE Id_documento = '"+str(idArchivo[0])+"'")
    
        banco = cur.fetchall()
                        
        banco.insert(0, ("Cuenta", "Fecha", "Referencia_num", "Importe", "Transaccion", "Leyenda_1", "Leyenda_2", "info_adicional", "Conciliado"))
        for datos in banco:
            hoja.append(datos)    
        
        hoja = archivo[nomHojas[2]]
            
        cur.execute("SELECT Num, Folio, Total, Acreditacion, Fecha_registro, Desde_Envios, Esta_En_Banco FROM acreditaciones WHERE Id_documento = '"+str(idArchivo[0])+"'")#OR Esta_En_Banco = 'No'
    
        acreditado = cur.fetchall()
                        
        acreditado.insert(0, ("Num", "Folio", "Total", "Acreditacion", "Fecha_registro", "En envios", "En Bancos"))
        for datos in acreditado:
            hoja.append(datos)
            
        hoja = archivo[nomHojas[3]]
               
        cur.execute("SELECT Tipo_trans, Descripcion, Nat, Order_id, Store_id, Store_name, Time_in, Monto, Auth_code, Marcar, Q6, Ultimos_dig, porcentaje_com, Comision, IVA_comision, Pago, Fecha, Hora, Naturaleza, ID_bem, D, Folio_id, Desde_Envios, Esta_En_Banco FROM netpay WHERE Id_documento = '"+str(idArchivo[0])+"'")#OR Esta_En_Banco = 'No'
    
        acreditado = cur.fetchall()
                        
        acreditado.insert(0, ("Tipo_trans", "Descripcion", "Nat", "Order_id", "Store_id", "Store_name", "Time_in", "Monto", "Auth_code", "Marcar", "Q6", "Ultimos_dig", "porcentaje_com", "Comision", "IVA_comision", "Pago", "Fecha", "Hora", "Naturaleza", "ID_bem", "D", "Folio_id", "En envios", "En Bancos"))
        for datos in acreditado:
            hoja.append(datos)
        
        archivo.save(rutaActual + "/Archivos_Conciliados/" + nombreArchivo+"_Conciliado.xlsx")
        
        noEncontrados(cur, nomHojas)
    else:
        borrarDeBD(cur, idArchivo)
  
def noEncontrados(cur, nomHojas):
    archivo = Workbook()
    
    hoja = archivo.active
    hoja.title = nomHojas[0]
        
    for i in nomHojas:
        if i != hoja.title:
            hoja = archivo.create_sheet(i)
    
    hoja = archivo[nomHojas[0]]        
    
    cur.execute("SELECT ID, F_trabajo, Fecha_alta, Sucursal, Caja, Turno, ID_envio, Sucursal_ref, Caja_ref, Tipo, Concepto, Detalle_concepto, ID_ref, Estado, Total, Referencia, Mov_banco, Mov_suc, Conciliacion FROM envios WHERE Conciliacion != 'Encontrado'") #WHERE Encontrado != 'Si'")#Encontrado = 'No'")

    envios = cur.fetchall()
                    
    envios.insert(0, ("ID Base de Datos", "F_trabajo", "Fecha_alta", "Sucursal", "Caja", "Turno", "ID_envio", "Sucursal_ref", "Caja_ref", 
                "Tipo", "Concepto", "Detalle_concepto", "ID_ref", "Estado", "Total", "Referencia", "Mov_banco", "Mov_suc", "Conciliado"))
    for datos in envios:
        hoja.append(datos)
    
    hoja = archivo[nomHojas[1]]
        
    cur.execute("SELECT ID, Cuenta, Fecha, Referencia_num, Importe, Transaccion, Leyenda_1, Leyenda_2, info_adicional, Conciliacion FROM banco WHERE Cuenta IS NOT NULL AND Conciliacion = 'No Encontrado'")# WHERE Conciliado = 'No'")

    banco = cur.fetchall()
                    
    banco.insert(0, ("ID Base de Datos", "Cuenta", "Fecha", "Referencia_num", "Importe", "Transaccion", "Leyenda_1", "Leyenda_2", "info_adicional", "Conciliado"))
    for datos in banco:
        hoja.append(datos)    
    
    hoja = archivo[nomHojas[2]]
        
    cur.execute("SELECT ID, Num, Folio, Total, Acreditacion, Fecha_registro, Desde_Envios, Esta_En_Banco FROM acreditaciones WHERE Folio IS NOT NULL AND Desde_Envios = 'No Encontrado' OR Esta_En_Banco = 'No Encontrado'")# WHERE Desde_Envios = 'No' OR Esta_En_Banco = 'No'")#OR Esta_En_Banco = 'No'

    acreditado = cur.fetchall()
                    
    acreditado.insert(0, ("ID Base de Datos", "Num", "Folio", "Total", "Acreditacion", "Fecha_registro", "En Envios", "En Bancos"))
    for datos in acreditado:
        hoja.append(datos)
        
    hoja = archivo[nomHojas[3]]
           
    cur.execute("SELECT ID, Tipo_trans, Descripcion, Nat, Order_id, Store_id, Store_name, Time_in, Monto, Auth_code, Marcar, Q6, Ultimos_dig, porcentaje_com, Comision, IVA_comision, Pago, Fecha, Hora, Naturaleza, ID_bem, D, Folio_id, Desde_Envios, Esta_En_Banco FROM netpay WHERE Monto IS NOT NULL AND (Desde_Envios = 'No Encontrado' OR Desde_Envios IS NULL OR Esta_En_Banco = 'No Encontrado')")# WHERE Desde_Envios = 'No' OR Desde_Envios IS NULL OR Esta_En_Banco = 'No'")#OR Esta_En_Banco = 'No'

    acreditado = cur.fetchall()
                    
    acreditado.insert(0, ("ID Base de Datos", "Tipo_trans", "Descripcion", "Nat", "Order_id", "Store_id", "Store_name", "Time_in", "Monto", "Auth_code", "Marcar", "Q6", "Ultimos_dig", "porcentaje_com", "Comision", "IVA_comision", "Pago", "Fecha", "Hora", "Naturaleza", "ID_bem", "D", "Folio_id", "En Envios","En Bancos"))
    for datos in acreditado:
        hoja.append(datos)
    
    archivo.save(rutaActual + "/Archivos_Conciliados/No_Encontrados.xlsx")
  
def crearDB(cursor):
    # Create table
    cur.execute('''CREATE TABLE IF NOT EXISTS archivos (ID integer PRIMARY KEY AUTOINCREMENT, Nombre text)''')
    
    cur.execute('''CREATE TABLE IF NOT EXISTS envios (ID integer PRIMARY KEY AUTOINCREMENT, F_trabajo real, Fecha_alta real, 
                Sucursal text, Caja text, Turno int, ID_envio int, Sucursal_ref text, Caja_ref text, Tipo text, Concepto text, 
                Detalle_concepto text, ID_ref int, Estado text, Total real, Referencia text, Mov_banco int, Mov_suc int, Id_documento integer, Id_Banco integer, Id_Netpay integer, Conciliacion text default "No Encontrado")''')#En_armstrong boolean default 0, En_netpay boolean default 0,
    
    cur.execute('''CREATE TABLE IF NOT EXISTS banco (ID integer PRIMARY KEY AUTOINCREMENT, Cuenta int, Fecha real, Referencia_num int, 
                Importe real, Transaccion text, Leyenda_1 text, Leyenda_2 text, info_adicional text, Id_agrupado_armstrong int, Id_agrupado_netpay int, Id_Envio integer, Id_documento integer, Conciliacion text default "No Encontrado")''')
    
    cur.execute('''CREATE TABLE IF NOT EXISTS acreditaciones (ID integer PRIMARY KEY AUTOINCREMENT, Num int, Folio text, Total real, 
                Acreditacion real, Fecha_registro real, Id_banco integer, Desde_Envios text default 'No Encontrado', Id_agrupado int, Id_documento integer, Esta_En_Banco text default "No Encontrado")''')#En_banco boolean default 0,
                          
    cur.execute('''CREATE TABLE IF NOT EXISTS netpay (ID integer PRIMARY KEY AUTOINCREMENT, Tipo_trans int, Descripcion text, Nat text, Order_id text, Store_id int,
                Store_name text, Time_in real, Monto real, Auth_code text, Marcar text, Q6 text, Ultimos_dig int,
                porcentaje_com real, Comision real, IVA_comision real, Pago real, Fecha real, Hora real,
                Naturaleza text, ID_bem text, D text, Folio_id text, Id_agrupado int, Id_documento integer, Id_banco integer, Id_Envio integer, Desde_Envios text default 'No Encontrado', Esta_En_Banco text default "No Encontrado")''')#En_banco bolean default 0,

def ingresarDatos(cur, arreglo, idArchivo):
    huboError = False
    seIngreso = True

    for i in range(len(arreglo)):
        if i == 0:
            for j in range(len(arreglo[i])):
                arreglo[i][j] = list(arreglo[i][j])
                arreglo[i][j].append(idArchivo[0])
    
            for j in range(len(arreglo[i])):
                arreglo[i][j] = tuple(arreglo[i][j])
                
            for datos in arreglo[i]:
                if datos[0] != "F. trabajo":
                    try:
                        cur.execute('''INSERT INTO envios(F_trabajo, Fecha_alta, Sucursal, Caja, Turno, ID_envio, Sucursal_ref, Caja_ref, 
                                        Tipo, Concepto, Detalle_concepto, ID_ref, Estado, Total, Referencia, Mov_banco, Mov_suc, Id_documento) 
                                        VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)''', datos)
                    except:
                        huboError = True    
                        seIngreso = False
            if huboError:
                print("ERROR, revise que no existan columnas o datos de mas en la pestaña 'ENVIOS'")
                break
            huboError = False
        if i == 1:
            for j in range(len(arreglo[i])):
                arreglo[i][j] = list(arreglo[i][j])
                arreglo[i][j].append(idArchivo[0])
    
            for j in range(len(arreglo[i])):
                arreglo[i][j] = tuple(arreglo[i][j])
                
            for datos in arreglo[i]:
                if datos[0] != "Cuenta":
                    try:
                        cur.execute('''INSERT INTO banco(Cuenta, Fecha, Referencia_num, Importe, Transaccion, 
                                    Leyenda_1, Leyenda_2, info_adicional, Id_documento) 
                                    VALUES (?,?,?,?,?,?,?,?,?)''', datos)
                    except:
                        huboError = True
                        seIngreso = False
            if huboError:
                print("ERROR, revise que no existan columnas o datos de mas en la pestaña 'BANCOS'")
                break
            huboError = False
        if i == 2:
            cur.execute('''SELECT Id_agrupado FROM acreditaciones ORDER BY Id_agrupado DESC LIMIT 1''')
            
            idAgrupado = cur.fetchone()
            
            try:
                idAgrupado = int(idAgrupado[0])
            except:
                idAgrupado = 0
            
            for j in range(len(arreglo[i])):
                arreglo[i][j] = list(arreglo[i][j])
                if str(arreglo[i][j][0]) == "No":
                    idAgrupado += 1

                arreglo[i][j].append(idAgrupado)
                arreglo[i][j].append(idArchivo[0])
    
            for j in range(len(arreglo[i])):
                arreglo[i][j] = tuple(arreglo[i][j])
                
            for datos in arreglo[i]:
                if type(datos[0]) != str: #type(datos[0]) == int
                    try:
                        cur.execute('''INSERT INTO acreditaciones(Num, Folio, Total, Acreditacion, Fecha_registro, Id_agrupado, Id_documento) 
                                VALUES (?,?,?,?,?,?,?)''', datos)    
                    except:
                        huboError = True
                        seIngreso = False
            if huboError:
                print("ERROR, revise que no existan columnas o datos de mas en la pestaña 'ACREDITACIONES'")
                break
            huboError = False
                                
        if i == 3:
            cur.execute('''SELECT Id_agrupado FROM netpay ORDER BY Id_agrupado DESC LIMIT 1''')
            
            idAgrupadoNet = cur.fetchone()
            
            try:
                idAgrupadoNet = int(idAgrupadoNet[0])
            except:
                idAgrupadoNet = 0
            
            #auxID = 0
            for j in range(len(arreglo[i])):
                #print(arreglo[i][j])
                arreglo[i][j] = list(arreglo[i][j])
                '''if idAgrupadoNet == 0 or type(arreglo[i][j][1]) != str or arreglo[i][j][1] == "Descripcion":
                    idAgrupadoNet += 1
                    auxID += 1
                    
                if auxID > 1:
                    idAgrupadoNet -= 1
                    auxID = 0'''
                    
                if arreglo[i][j][1] == "Descripcion" or idAgrupadoNet == 0:
                    idAgrupadoNet += 1
                
                elif j > 0:
                    if type(arreglo[i][j][1]) == str and type(arreglo[i][j-1][1]) != str:
                        idAgrupadoNet += 1
                 
                arreglo[i][j].append(idAgrupadoNet)
                arreglo[i][j].append(idArchivo[0])
            
            for j in range(len(arreglo[i])):
                arreglo[i][j] = tuple(arreglo[i][j])
                
            for datos in arreglo[i]:
                if datos[1] != "Descripcion": #type(datos[1]) == str
                    try:
                        cur.execute('''INSERT INTO netpay(Tipo_trans, Descripcion, Nat, Order_id, Store_id,
                                    Store_name, Time_in, Monto, Auth_code, Marcar, Q6, Ultimos_dig, porcentaje_com, 
                                    Comision, IVA_comision, Pago, Fecha, Hora, Naturaleza, ID_bem, D, Folio_id, Id_agrupado, Id_documento)  
                                    VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)''', datos)
                    except:
                        huboError = True
                        seIngreso = False
            if huboError:
                print("ERROR, revise que no existan columnas o datos de mas en la pestaña 'NET PAY'")
                break
            huboError = False
    return seIngreso
        
def enviosArmstrong(cur):
    #En_armstrong = 1,
    cur.execute('''UPDATE envios SET Conciliacion = "Encontrado" WHERE Referencia IN(
                SELECT envios.Referencia from envios LEFT JOIN acreditaciones 
                ON envios.Referencia = substr(acreditaciones.Folio,1,8) 
                WHERE acreditaciones.Folio IS NOT NULL AND envios.Detalle_concepto = "Armstrong" AND acreditaciones.Total = envios.Total)''')
    
    cur.execute('''UPDATE acreditaciones SET Desde_Envios = 'Encontrado' WHERE Folio IN(
                SELECT acreditaciones.Folio from acreditaciones LEFT JOIN envios 
                ON substr(acreditaciones.Folio,1,8) = envios.Referencia
                WHERE acreditaciones.Folio IS NOT NULL AND envios.Detalle_concepto = "Armstrong" AND acreditaciones.Total = envios.Total)''')#(SELECT sum(total) FROM envios Where Detalle_concepto = 'Armstrong' GROUP BY Referencia)
 
    cur.execute("UPDATE acreditaciones SET Desde_Envios = NULL WHERE Folio IS NULL")
    
    cur.execute('''UPDATE envios SET Conciliacion = "Monto diferente acreditado" WHERE Referencia IN(
                SELECT envios.Referencia from envios LEFT JOIN acreditaciones 
                ON envios.Referencia = substr(acreditaciones.Folio,1,8) 
                WHERE acreditaciones.Folio IS NOT NULL AND envios.Detalle_concepto = "Armstrong" AND acreditaciones.Total != envios.Total AND Conciliacion != 'Encontrado')''')
    
    cur.execute('''UPDATE envios SET Conciliacion = "Referencia mal registrada" WHERE length(Referencia) != 8 AND Detalle_Concepto = "Armstrong" AND Conciliacion = "No Encontrado"''')
    
def enviosNetpay(cur):
    cur.execute('''UPDATE netpay SET Auth_code = NULL WHERE Auth_code = "' "''')
    cur.execute("SELECT Id, Monto, substr(Auth_code,2) FROM netpay WHERE Desde_Envios = 'No Encontrado' AND Auth_code IS NOT NULL")
    
    netpay = cur.fetchall()
    
    cur.execute("SELECT Id, Total, Referencia FROM envios WHERE Detalle_concepto = 'NETPAY' AND Conciliacion = 'No Encontrado'")
    
    envios = cur.fetchall()
    
    for i in range(len(netpay)):
         netpay[i] = list(netpay[i])
         netpay[i].append(0)

    for i in range(len(envios)):
         envios[i] = list(envios[i])
         envios[i].append(0)    
    
    for i in range(len(envios)):
        for j in range(len(netpay)):
            if envios[i][1] == netpay[j][1] and netpay[j][2] in envios[i][2] and envios[i][3] == 0 and netpay[j][3] == 0:
                envios[i][3] = 1
                netpay[j][3] = 1
                cur.execute("UPDATE netpay SET Desde_Envios = 'Encontrado', Id_Envio = '"+str(envios[i][0])+"' WHERE Id = '"+str(netpay[j][0])+"'")
                cur.execute("UPDATE envios SET Conciliacion = 'Encontrado', Id_Netpay = '"+str(netpay[j][0])+"' WHERE Id = '"+str(envios[i][0])+"'")
                #print(envios[i],'=',netpay[j])  
                break
                
def armstrongBanco(cur):
    cur.execute("SELECT DISTINCT Id_agrupado, acreditacion FROM acreditaciones WHERE Esta_En_Banco = 'No Encontrado'")
    acredita = cur.fetchall()

    for i in (acredita):
        idAgrupado = (i[0]),
        acreditaArmstrong = (i[1]),
        cur.execute("SELECT Id FROM banco WHERE Conciliacion = 'No Encontrado' AND Cuenta = 586082157 AND Transaccion LIKE '%GUADALAJARA' AND Importe = ? LIMIT 1", acreditaArmstrong)
        
        idBanco = cur.fetchall()
        if len(idBanco) != 0:
            idBanco = (idBanco[0][0]),
            #En_banco = 1,
            cur.execute("UPDATE acreditaciones SET Esta_En_Banco = 'Encontrado', Id_banco = '"+str(idBanco[0])+"' WHERE Id_agrupado = '"+str(i[0])+"'")

            #En_armstrong = 1,
            cur.execute("UPDATE banco SET Conciliacion = 'Encontrado', Id_agrupado_armstrong = '"+str(idAgrupado[0])+"' WHERE id = '"+str(idBanco[0])+"' AND Cuenta IS NOT NULL")
  
    cur.execute("UPDATE acreditaciones SET Esta_En_Banco = NULL WHERE Total IS NULL")   
    cur.execute("UPDATE banco Set Conciliacion = NULL WHERE Cuenta IS NULL")            

def netpayBanco(cur):
    #cur.execute('''DELETE FROM netpay WHERE Descripcion IS NULL''')
    cur.execute("SELECT DISTINCT Id_agrupado FROM netpay WHERE Esta_En_Banco = 'No Encontrado'")
    idAgrupado = cur.fetchall()
    
    for i in idAgrupado:
        cur.execute("SELECT ROUND(SUM(Monto),2) FROM netpay WHERE Id_agrupado = '"+str(i[0])+"'")
        
        total = cur.fetchone()
        
        #print(i, total)
        
        cur.execute("SELECT Id FROM banco WHERE Conciliacion = 'No Encontrado' AND Cuenta = 1005694868 AND Leyenda_1 LIKE 'BN%' AND Importe = '"+str(total[0])+"' LIMIT 1")
        
        idBanco = cur.fetchone()
        if type(idBanco) == tuple:
            #En_banco = 1,
            cur.execute("UPDATE netpay SET Esta_En_Banco = 'Encontrado', Id_banco = '"+str(idBanco[0])+"' WHERE Id_agrupado = '"+str(i[0])+"'")

            #En_netpay = 1,
            cur.execute("UPDATE banco SET Conciliacion = 'Encontrado', Id_agrupado_netpay = '"+str(i[0])+"' WHERE id = '"+str(idBanco[0])+"'")       
            
    cur.execute("UPDATE netpay SET Esta_En_Banco = NULL WHERE Monto IS NULL")    
    cur.execute("UPDATE netpay SET Desde_Envios = NULL WHERE Monto IS NULL OR Nat = 'C'")    
    cur.execute("UPDATE banco Set Conciliacion = NULL WHERE Cuenta IS NULL") 

def enviosBanco(cur):
    cur.execute("""SELECT Id, UPPER(Sucursal), Total FROM envios WHERE Detalle_Concepto NOT IN ('Armstrong', 'NETPAY') AND Conciliacion = 'No Encontrado'""")
    
    datosEnvios = cur.fetchall()
    
    cur.execute("""SELECT Id, CASE Referencia_num WHEN 117 THEN 'CHICALOTE' WHEN 216 THEN 'AVENIDA' WHEN 315 THEN 'NUEZ 1' WHEN 414 THEN 'NUEZ 3' WHEN 521 THEN 'TRIGO' END Referencia, Importe FROM banco WHERE Referencia IN('CHICALOTE', 'AVENIDA', 'NUEZ 1', 'NUEZ 3', 'TRIGO') AND Conciliacion = 'No Encontrado'""")
    
    datosBanco = cur.fetchall()
    
    for i in range(len(datosEnvios)):
        datosEnvios[i] = list(datosEnvios[i])
        datosEnvios[i].append(0)
    
    for i in range(len(datosBanco)):
        datosBanco[i] = list(datosBanco[i])
        datosBanco[i].append(0)

    for i in range(len(datosEnvios)):
        for j in range(len(datosBanco)):
            if datosEnvios[i][1] == datosBanco[j][1] and datosEnvios[i][2] == datosBanco[j][2] and datosEnvios[i][3] == 0 and datosBanco[j][3] == 0:
                datosEnvios[i][3] = 1
                datosBanco[j][3] = 1
                cur.execute("UPDATE banco SET Conciliacion = 'Encontrado', Id_Envio = '"+str(datosEnvios[i][0])+"' WHERE ID = '"+str(datosBanco[j][0])+"'")
                cur.execute("UPDATE envios SET Conciliacion = 'Encontrado', Id_Banco = '"+str(datosBanco[j][0])+"' WHERE ID = '"+str(datosEnvios[i][0])+"'")
                #print(datosEnvios[i],' = ',datosBanco[j])
                break
    
    '''for i in datosBanco:
        print(type(i[2]),i[2], i[3])
    print("\n")
    for i in datosEnvios:
        print(i[0],type(i[2]),i[2], i[3])'''
    
rutaActual = os.getcwd()

if not os.path.exists(rutaActual + "/Archivos_Conciliados"):
    os.mkdir(rutaActual + "/Archivos_Conciliados")   
    
if not os.path.exists(rutaActual + "/No_Borrar_BD"):
    os.mkdir(rutaActual + "/No_Borrar_BD")   
    #hide_file("No_Borrar")

ejecutar = True
    
while ejecutar:
        
    #inicio del programa donde se ingresa el nombre del archivo excel a conciliar
    while True:
        
        time.sleep(1.8)
        system("cls")
        
        print("Conciliacion de depositos\n")
        nombreArchivo = input("Ingrese el nombre del archivo o '0' para salir: ").upper()
        
        if nombreArchivo != '0': 
            yaExiste = False
            
            #establecemos conexion con la BD y en caso de que no exista se crea
            con = sqlite3.connect(rutaActual + "/No_Borrar_BD/BD_conciliacion.db")
            
            #se crea una instancia de la bd con un cursor para poder manejar la bd
            cur = con.cursor()
            
            crearDB(cur)
            
            cur.execute("SELECT Nombre FROM archivos")
            
            archivoBD = cur.fetchall()
            
            for nombre in archivoBD:
                if nombre[0] == nombreArchivo:
                    yaExiste = True
                    break
        
            if yaExiste:
                con.close()
                print("ERROR, ya se concilio un archivo con ese nombre anteriormente")
            else:
                try:
                    libro = load_workbook(filename = rutaActual +'/'+ nombreArchivo + ".xlsx")
                    cur.execute("INSERT INTO archivos(Nombre) VALUES('"+nombreArchivo+"')")
                    print("Conciliando el archivo '"+nombreArchivo+"'")
                    
                    time.sleep(1.2)
                    nombreHojas = libro.sheetnames
        
                    hojas = []
                    arreglo = []
                    
                    for i in range(len(libro.sheetnames)):
                        hojas.append(libro[nombreHojas[i]])
                        arreglo.append([])
                        if i == 0:
                            #max_col=hojas[i].max_column
                            for row in hojas[i].iter_rows(min_row=1, max_col = 17, max_row = hojas[i].max_row, values_only=True):
                                arreglo[i].append(row)
                        if i == 1:
                            for row in hojas[i].iter_rows(min_row=1, max_col = 8, max_row = hojas[i].max_row, values_only=True):
                                arreglo[i].append(row)
                        if i == 2:
                            for row in hojas[i].iter_rows(min_row=1, max_col = 5, max_row = hojas[i].max_row, values_only=True):
                                arreglo[i].append(row)
                        if i == 3:
                            for row in hojas[i].iter_rows(min_row=1, max_col = 22, max_row = hojas[i].max_row, values_only=True):
                                arreglo[i].append(row)
                        #print(len(arreglo[i][0]))
                    
                    cur.execute("SELECT ID FROM archivos WHERE Nombre = '"+nombreArchivo+"'")
                    idArchivo = cur.fetchone()
                        
                    ingresado = ingresarDatos(cur, arreglo, idArchivo)
                    
                    enviosArmstrong(cur)
                    
                    enviosNetpay(cur)
                    
                    armstrongBanco(cur)
                    
                    netpayBanco(cur)
                    
                    enviosBanco(cur)
                    
                    crearArchivo(cur, nombreHojas, nombreArchivo, idArchivo, ingresado)
                    
                    # Save (commit) the changes
                    con.commit()
                    
                    # We can also close the connection if we are done with it.
                    # Just be sure any changes have been committed or they will be lost.
                    con.close()
                    
                    libro.save(rutaActual +'/'+ nombreArchivo + ".xlsx")
                except:
                    print("ERROR, el archivo "+ nombreArchivo +" no existe") 
        else:
            ejecutar = False
            break
            
print("\nCerrando...")
time.sleep(2.5)